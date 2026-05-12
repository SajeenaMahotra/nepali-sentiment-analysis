"""Build XLSX workbook + README for the 100K marketplace v2 dataset."""
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

OUT = Path("/home/claude/marketplace_v2")
df = pd.read_csv(OUT / "kathmandu_marketplace_reviews_100k.csv")

# ---- styling --------------------------------------------------------------
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="Arial", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row=1):
    for c in ws[row]:
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER


def autosize(ws, max_w=55):
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        m = 0
        for c in col_cells:
            if c.value is None: continue
            l = min(len(str(c.value)), max_w)
            if l > m: m = l
        ws.column_dimensions[letter].width = max(10, min(m+2, max_w))


# ---- workbook -------------------------------------------------------------
wb = Workbook()

# README sheet
ws = wb.active; ws.title = "README"
ws.append(["Field","Description"]); style_header(ws)
readme_rows = [
    ["Dataset",         "Kathmandu Marketplace Code-Mixed Reviews v2 (100K)"],
    ["Version",         "2.0.0"],
    ["Created",         datetime.now().strftime("%Y-%m-%d")],
    ["Project",         "Sentiment Analysis of Code-Mixed Nepali-English Product Reviews"],
    ["Locale",          "Generic Kathmandu-valley online marketplace (no real company named)"],
    ["Total reviews",   f"{len(df):,}"],
    ["Class split",     "Positive 58,000 / Negative 30,000 / Neutral 12,000 (~58/30/12 — Pradhananga & Sah-style natural skew)"],
    ["Unique texts",    "83,030 / 100,000 = 83.0% unique (low duplication for healthy ML training)"],
    ["Date range",      f"{df['review_date'].min()} → {df['review_date'].max()}"],
    ["Schema",          "22 columns — IDENTICAL to v1, drop-in replacement"],
    ["Splits",          "Stratified 70/15/15 by sentiment — see split_*.csv"],
    ["Reproducibility", "random.seed(42) — re-running generate_dataset.py reproduces every row"],
    ["Diversity",       "183 SKUs across 10 categories; 5 fragment slots per sentiment + slot-based phrase generators"],
    ["License",         "Synthetic — free for thesis, code repos, derived works"],
]
for r in readme_rows: ws.append(r)
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 110
for row in ws.iter_rows(min_row=2):
    for c in row: c.font = BODY_FONT; c.alignment = WRAP

# What's new vs v1
ws = wb.create_sheet("What's New (v1 vs v2)")
ws.append(["Aspect","v1 (2,500 rows)","v2 (100,000 rows)"]); style_header(ws)
diffs = [
    ["Total reviews",          "2,500",  "100,000"],
    ["Product catalogue size", "~46 SKUs","183 SKUs"],
    ["Categories",             "8",      "10 (added Sports & Outdoor; expanded all)"],
    ["Sellers",                "20",     "30"],
    ["Locations",              "15",     "20"],
    ["Payment methods",        "6",      "7 (added FonePay)"],
    ["Delivery partners",      "5",      "8"],
    ["Fragment slots",         "3 (opener, body, closer)", "5 (opener, product_aspect, delivery_aspect, service_aspect, closer)"],
    ["Templates per slot",     "~20",    "~30-50"],
    ["Slot-based phrase gen",  "No",     "Yes — {aspect} × {quality} × {intensifier} explodes combinatorially"],
    ["Mixed-aspect reviews",   "No",     "Yes (~8% of positives include a delivery gripe; ~6% of negatives include positive delivery)"],
    ["Short reviews",          "No",     "Yes — 10% are single-phrase reviews like 'Mast cha!' / 'Bekar product.'"],
    ["Schema",                 "22 cols","22 cols (identical — drop-in replacement)"],
    ["Unique text rate",       "~95% (small sample)", "83% (healthy at scale)"],
]
for r in diffs: ws.append(r)
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 70
for row in ws.iter_rows(min_row=2):
    for c in row: c.font = BODY_FONT; c.alignment = WRAP

# Data dictionary (same 22 columns as v1)
ws = wb.create_sheet("Data Dictionary")
ws.append(["Column","Type","Description"]); style_header(ws)
DD = [
    ("review_id",            "string",    "Primary key. e.g. R000001."),
    ("product_id",           "string",    "Stable ID for the product (hash of product name)."),
    ("product_name",         "string",    "Full product name with model/variant."),
    ("product_category",     "string",    "One of 10 categories (Mobiles, Fashion, Beauty, etc.)."),
    ("brand",                "string",    "Brand name."),
    ("product_price_npr",    "int",       "Listed price in Nepalese Rupees."),
    ("seller_name",          "string",    "Generic Kathmandu seller name."),
    ("delivery_partner",     "string",    "Pathao, Aramex, Daraz Express, etc."),
    ("reviewer_location",    "string",    "City — weighted toward Kathmandu valley."),
    ("payment_method",       "string",    "COD / eSewa / Khalti / IME Pay / Card / ConnectIPS / FonePay."),
    ("review_date",          "datetime",  "When the review was posted (Jan 2024 → Apr 2026)."),
    ("review_text",          "string",    "**The label target — Romanized Nepali + English code-mixed.**"),
    ("rating",               "int 1-5",   "Star rating (correlated with sentiment_label)."),
    ("sentiment_label",      "string",    "**Primary label**: positive / negative / neutral."),
    ("verified_purchase",    "bool",      "Whether the review is from a verified buyer."),
    ("helpful_count",        "int",       "Number of users who marked the review helpful."),
    ("review_length_chars",  "int",       "Character count of review_text."),
    ("review_length_tokens", "int",       "Whitespace-split token count."),
    ("english_token_ratio",  "float",     "Fraction of tokens matching English hint vocabulary."),
    ("nepali_token_ratio",   "float",     "Fraction of tokens matching Romanized Nepali hint vocabulary."),
    ("language_dominance",   "string",    "Nepali-dominant / English-dominant / Balanced (derived)."),
    ("has_emoji",            "bool",      "Whether the review_text contains any emoji codepoint."),
]
for r in DD: ws.append(r)
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 90
for row in ws.iter_rows(min_row=2):
    for c in row: c.font = BODY_FONT; c.alignment = WRAP

# Summary stats
ws = wb.create_sheet("Summary Stats")
ws.append(["Metric","Value"]); style_header(ws)
stats = [
    ["Total reviews", len(df)],
    ["Unique review texts", df['review_text'].nunique()],
    ["Unique products", df['product_name'].nunique()],
    ["Unique sellers", df['seller_name'].nunique()],
    ["Avg review length (chars)", round(df['review_length_chars'].mean(), 1)],
    ["Median review length (chars)", int(df['review_length_chars'].median())],
    ["Avg review length (tokens)", round(df['review_length_tokens'].mean(), 1)],
    ["Reviews with emoji %", round(100*df['has_emoji'].mean(), 2)],
    ["Verified purchase %", round(100*df['verified_purchase'].mean(), 2)],
    ["Avg helpful_count", round(df['helpful_count'].mean(), 1)],
]
for r in stats: ws.append(r)

ws.append([]); ws.append(["Sentiment distribution"])
for k, v in df['sentiment_label'].value_counts().items(): ws.append([k, int(v)])

ws.append([]); ws.append(["Language dominance"])
for k, v in df['language_dominance'].value_counts().items(): ws.append([k, int(v)])

ws.append([]); ws.append(["Rating distribution"])
for k, v in df['rating'].value_counts().sort_index().items(): ws.append([int(k), int(v)])

ws.append([]); ws.append(["Top categories"])
for k, v in df['product_category'].value_counts().items(): ws.append([k, int(v)])

ws.append([]); ws.append(["Payment methods"])
for k, v in df['payment_method'].value_counts().items(): ws.append([k, int(v)])

autosize(ws)
for row in ws.iter_rows(min_row=2):
    for c in row:
        if c.font != HEADER_FONT: c.font = BODY_FONT


# Sample sheets — 1000 rows each (XLSX can't hold 100K rows comfortably)
def sample_sheet(name, sub, n=1000):
    ws = wb.create_sheet(name)
    for r in dataframe_to_rows(sub.head(n), index=False, header=True):
        ws.append(r)
    style_header(ws)
    ws.freeze_panes = "A2"
    autosize(ws, max_w=60)


sample_sheet("Positive Sample", df[df.sentiment_label == "positive"], 1000)
sample_sheet("Negative Sample", df[df.sentiment_label == "negative"], 1000)
sample_sheet("Neutral Sample",  df[df.sentiment_label == "neutral"],  500)
sample_sheet("Random Mix Sample", df.sample(1000, random_state=1).reset_index(drop=True), 1000)

xlsx_path = OUT / "kathmandu_marketplace_reviews_100k.xlsx"
wb.save(xlsx_path)
print("Wrote", xlsx_path)
print("Sheets:", wb.sheetnames)
